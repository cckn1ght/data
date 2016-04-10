from openpyxl import load_workbook
from pymongo import MongoClient

class mongo_insertor(object):
    def __init__(self, db_name, db_collection):
        client = MongoClient('localhost', 27017)
        db = client[db_name]
        self.collection = db[db_collection]

    def insert(self):
        collection = self.collection
        for i in range(1, 7):
            name = 'user_data_' + str(i) + '.xlsx'
            wb = load_workbook(name)
            # print(wb.get_sheet_names())
            sheet = wb.active
            # rows = sheet.get_highest_row()
            # print(rows)
            for row in sheet.iter_rows(row_offset=1):
                if row[0].value is None:
                    continue
                content = {}
                content['ID'] = row[0].value
                content['screen_name'] = '@' + str(row[1].value)
                content['name'] = row[2].value
                content['location'] = row[3].value
                content['description'] = row[4].value
                content['followers_count'] = int(row[6].value)
                content['friends_count'] = int(row[7].value)
                content['statuses_count'] = int(row[8].value)
                content['listed_count'] = int(row[9].value)
                content['is_protected'] = row[12].value
                content['verified'] = row[13].value
                content['member_since'] = row[16].value
                # matching_item = collection.find_one(
                #     {'screen_name': content['screen_name'], })
                # if matching_item is None:
                collection.insert_one(content)
        print('insertion done!')
def main():
    db_name = 'nurse_users_2'
    db_collection = 'users'
    insertor1 = mongo_insertor(db_name, db_collection)
    insertor1.insert()

if __name__ == "__main__":
    main()
